"""
生产管理系统 ProductionSystem
"""
import sqlite3, os, sys, webbrowser, tempfile, json, shutil, calendar
import hashlib, secrets, base64
from datetime import date, timedelta
from tkinter import *
from tkinter import ttk, messagebox

def _rp(rel):
    try: base = sys._MEIPASS
    except: base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, rel)

if getattr(sys, "frozen", False): _BASE = os.path.dirname(sys.executable)
else: _BASE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(_BASE, 'data.db')
ECHARTS_PATH = _rp('echarts.min.js')
YEN = chr(0xa5)
LOGIN_CRED_PATH = os.path.join(_BASE, '.login_cred')

ALL_PERMS = ['record_add','record_delete','record_edit','material_manage','worker_manage',
             'process_manage','assignment_manage','chart_view','summary_view','export_excel','user_manage']
ROLE_ADMIN = 'admin'
ROLE_LEADER = 'leader'
ROLE_WORKER = 'worker'

BG = '#f0f2f5'; CARD = '#ffffff'; PRIMARY = '#1a73e8'
ACCENT = '#e67e22'; GREEN = '#27ae60'; RED = '#e74c3c'; DARK = '#2c3e50'

def _set_placeholder(entry, text):
    entry._ph_text = text
    entry.insert(0, text)
    entry.config(fg='#999999')
    def _in(e):
        if entry.get() == entry._ph_text:
            entry.delete(0, END); entry.config(fg='#333333')
    def _out(e):
        if not entry.get().strip():
            entry.insert(0, entry._ph_text); entry.config(fg='#999999')
    entry.bind('<FocusIn>', _in, '+'); entry.bind('<FocusOut>', _out, '+')

def _hash_pw(pw):
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', pw.encode(), salt.encode(), 100000).hex()
    return salt + h

def _verify_pw(pw, stored):
    salt = stored[:32]
    h = hashlib.pbkdf2_hmac('sha256', pw.encode(), salt.encode(), 100000).hex()
    return (salt + h) == stored

def _save_cred(un, pw):
    try:
        data = json.dumps({'u':un,'p':base64.b64encode(pw.encode()).decode()})
        with open(LOGIN_CRED_PATH, 'w') as f: f.write(data)
    except: pass

def _load_cred():
    try:
        d = json.loads(open(LOGIN_CRED_PATH).read())
        d['p'] = base64.b64decode(d['p']).decode()
        return d
    except: return None

def _clear_cred():
    try:
        if os.path.exists(LOGIN_CRED_PATH): os.remove(LOGIN_CRED_PATH)
    except: pass

# ── Database ──

def init_db():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS materials (id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT NOT NULL UNIQUE,price REAL DEFAULT 0)")
    try: c.execute("ALTER TABLE materials ADD COLUMN price REAL DEFAULT 0")
    except: pass
    c.execute("UPDATE materials SET price=5.0 WHERE price IS NULL")
    c.execute("CREATE TABLE IF NOT EXISTS workers (id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT NOT NULL UNIQUE,group_name TEXT NOT NULL DEFAULT '')")
    try: c.execute("ALTER TABLE workers ADD COLUMN group_name TEXT NOT NULL DEFAULT ''")
    except: pass
    c.execute("CREATE TABLE IF NOT EXISTS processes (id INTEGER PRIMARY KEY AUTOINCREMENT,material TEXT NOT NULL,process_name TEXT NOT NULL,unit_price REAL NOT NULL DEFAULT 0,UNIQUE(material,process_name))")
    c.execute("CREATE TABLE IF NOT EXISTS worker_processes (id INTEGER PRIMARY KEY AUTOINCREMENT,worker_id INTEGER NOT NULL,process_id INTEGER NOT NULL,UNIQUE(worker_id,process_id))")
    c.execute("CREATE TABLE IF NOT EXISTS records (id INTEGER PRIMARY KEY AUTOINCREMENT,worker_id INTEGER NOT NULL DEFAULT 0,process_id INTEGER NOT NULL DEFAULT 0,quantity REAL NOT NULL,unit_price REAL NOT NULL,record_date TEXT NOT NULL,created_at TEXT DEFAULT (datetime('now','localtime')))")
    try: c.execute("ALTER TABLE records ADD COLUMN worker_id INTEGER")
    except: pass
    c.execute("CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY AUTOINCREMENT,username TEXT NOT NULL UNIQUE,password_hash TEXT NOT NULL,display_name TEXT NOT NULL DEFAULT '',role TEXT NOT NULL DEFAULT 'worker',worker_id INTEGER DEFAULT 0,created_at TEXT DEFAULT (datetime('now','localtime')))")
    try: c.execute("ALTER TABLE users ADD COLUMN worker_id INTEGER DEFAULT 0")
    except: pass
    c.execute("CREATE TABLE IF NOT EXISTS user_permissions (id INTEGER PRIMARY KEY AUTOINCREMENT,username TEXT NOT NULL,perm_key TEXT NOT NULL,allowed INTEGER NOT NULL DEFAULT 0,UNIQUE(username,perm_key))")
    c.execute("SELECT COUNT(*) FROM users")
    if c.fetchone()[0] == 0:
        pw = _hash_pw('admin123')
        c.execute("INSERT INTO users (username,password_hash,display_name,role) VALUES (?,?,?,?)", ('admin',pw,'系统管理员','admin'))
        c.execute("INSERT INTO workers (name,group_name) VALUES (?,?)", ('管理员','管理'))
        for pk in ALL_PERMS:
            c.execute("INSERT OR IGNORE INTO user_permissions (username,perm_key,allowed) VALUES (?,?,1)", ('admin',pk))
        c.execute("INSERT OR IGNORE INTO workers (name,group_name) VALUES ('管理员','管理')")
    conn.commit(); conn.close()

def get_conn():
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row; return conn

def get_users():
    conn = get_conn(); rows = conn.execute("SELECT * FROM users ORDER BY id").fetchall(); conn.close()
    return [dict(r) for r in rows]

def get_user_by_worker_id(wid):
    conn = get_conn(); r = conn.execute("SELECT * FROM users WHERE worker_id=?", (wid,)).fetchone(); conn.close()
    return dict(r) if r else None

def add_user(un, pw, dn='', role='worker', worker_id=0):
    conn = get_conn()
    try:
        h = _hash_pw(pw)
        conn.execute("INSERT INTO users (username,password_hash,display_name,role,worker_id) VALUES (?,?,?,?,?)", (un,h,dn,role,worker_id))
        for pk in ALL_PERMS:
            conn.execute("INSERT OR IGNORE INTO user_permissions (username,perm_key,allowed) VALUES (?,?,0)", (un,pk))
        conn.commit(); conn.close(); return True
    except: conn.close(); return False

def update_user_pw(un, pw):
    conn = get_conn(); h = _hash_pw(pw)
    conn.execute("UPDATE users SET password_hash=? WHERE username=?", (h,un))
    conn.commit(); conn.close()

def update_user_display(un, dn, role, worker_id=0):
    conn = get_conn()
    conn.execute("UPDATE users SET display_name=?,role=?,worker_id=? WHERE username=?", (dn,role,worker_id,un))
    conn.commit(); conn.close()

def delete_user(un):
    if un == 'admin': return False
    conn = get_conn()
    conn.execute("DELETE FROM users WHERE username=?", (un,))
    conn.execute("DELETE FROM user_permissions WHERE username=?", (un,))
    conn.commit(); conn.close(); return True

def get_user_perms(un):
    conn = get_conn()
    rows = conn.execute("SELECT perm_key,allowed FROM user_permissions WHERE username=?", (un,)).fetchall()
    conn.close()
    return {r['perm_key']:r['allowed'] for r in rows}

def set_user_perm(un, perm_key, allowed):
    conn = get_conn()
    conn.execute("INSERT OR REPLACE INTO user_permissions (username,perm_key,allowed) VALUES (?,?,?)", (un,perm_key,allowed))
    conn.commit(); conn.close()

def get_materials():
    conn = get_conn(); rows = conn.execute("SELECT * FROM materials ORDER BY name").fetchall(); conn.close()
    return [dict(r) for r in rows]

def add_material(name, price=0):
    conn = get_conn()
    try: conn.execute("INSERT INTO materials (name,price) VALUES (?,?)", (name,price)); conn.commit(); return True
    except: return False
    finally: conn.close()

def update_material(mid, name, price):
    conn = get_conn(); conn.execute("UPDATE materials SET name=?,price=? WHERE id=?", (name,price,mid))
    conn.commit(); conn.close()

def delete_material(mid):
    conn = get_conn(); conn.execute("DELETE FROM materials WHERE id=?", (mid,)); conn.commit(); conn.close()

def get_workers():
    conn = get_conn(); rows = conn.execute("SELECT * FROM workers ORDER BY name").fetchall(); conn.close()
    return [dict(r) for r in rows]

def add_worker(name, group=''):
    conn = get_conn()
    try: conn.execute("INSERT INTO workers (name,group_name) VALUES (?,?)", (name,group)); conn.commit(); return True
    except: return False
    finally: conn.close()

def update_worker(wid, name, group):
    conn = get_conn(); conn.execute("UPDATE workers SET name=?,group_name=? WHERE id=?", (name,group,wid))
    conn.commit(); conn.close()

def delete_worker(wid):
    conn = get_conn(); conn.execute("DELETE FROM workers WHERE id=?", (wid,))
    conn.execute("DELETE FROM worker_processes WHERE worker_id=?", (wid,))
    conn.commit(); conn.close()

def get_processes():
    conn = get_conn(); rows = conn.execute("SELECT p.*,m.price FROM processes p LEFT JOIN materials m ON p.material=m.name ORDER BY p.material,p.process_name").fetchall(); conn.close()
    return [dict(r) for r in rows]

def add_process(material, pname, unit_price):
    conn = get_conn()
    try: conn.execute("INSERT INTO processes (material,process_name,unit_price) VALUES (?,?,?)", (material,pname,unit_price)); conn.commit(); return True
    except: return False
    finally: conn.close()

def update_process(pid, material, pname, unit_price):
    conn = get_conn(); conn.execute("UPDATE processes SET material=?,process_name=?,unit_price=? WHERE id=?", (material,pname,unit_price,pid))
    conn.commit(); conn.close()

def delete_process(pid):
    conn = get_conn(); conn.execute("DELETE FROM processes WHERE id=?", (pid,))
    conn.execute("DELETE FROM worker_processes WHERE process_id=?", (pid,))
    conn.commit(); conn.close()

def get_worker_processes(wid):
    conn = get_conn(); rows = conn.execute("SELECT process_id FROM worker_processes WHERE worker_id=?", (wid,)).fetchall()
    conn.close()
    return [r['process_id'] for r in rows]

def assign_worker_process(wid, pid):
    conn = get_conn()
    try: conn.execute("INSERT INTO worker_processes (worker_id,process_id) VALUES (?,?)", (wid,pid)); conn.commit()
    except: pass
    finally: conn.close()

def unassign_worker_process(wid, pid):
    conn = get_conn(); conn.execute("DELETE FROM worker_processes WHERE worker_id=? AND process_id=?", (wid,pid))
    conn.commit(); conn.close()

def add_record(wid, pid, qty, price, d):
    conn = get_conn()
    try:
        conn.execute("INSERT INTO records (worker_id,process_id,quantity,unit_price,record_date) VALUES (?,?,?,?,?)", (wid,pid,qty,price,d))
        conn.commit(); return True
    except: return False
    finally: conn.close()

def update_record(rid, wid, pid, qty, price, d):
    conn = get_conn(); conn.execute("UPDATE records SET worker_id=?,process_id=?,quantity=?,unit_price=?,record_date=? WHERE id=?", (wid,pid,qty,price,d,rid))
    conn.commit(); conn.close()

def delete_record(rid):
    conn = get_conn(); conn.execute("DELETE FROM records WHERE id=?", (rid,)); conn.commit(); conn.close()

def get_all_records(user=None):
    conn = get_conn()
    if user and user.get('role') == 'worker' and user.get('worker_id'):
        rows = conn.execute("""
            SELECT r.*,w.name AS worker_name,w.group_name,p.material,p.process_name,p.unit_price AS default_price
            FROM records r LEFT JOIN workers w ON r.worker_id=w.id LEFT JOIN processes p ON r.process_id=p.id
            WHERE r.worker_id=? ORDER BY r.id DESC LIMIT 500
        """, (user['worker_id'],)).fetchall()
    else:
        rows = conn.execute("""
            SELECT r.*,w.name AS worker_name,w.group_name,p.material,p.process_name,p.unit_price AS default_price
            FROM records r LEFT JOIN workers w ON r.worker_id=w.id LEFT JOIN processes p ON r.process_id=p.id
            ORDER BY r.id DESC LIMIT 500
        """).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_stats(start_date=None, end_date=None, process_filter=None, worker_filter=None):
    conn = get_conn()
    wh = []; pa = []
    if start_date: wh.append("r.record_date >= ?"); pa.append(start_date)
    if end_date: wh.append("r.record_date <= ?"); pa.append(end_date)
    if process_filter: wh.append("r.process_id=?"); pa.append(process_filter)
    if worker_filter: wh.append("r.worker_id=?"); pa.append(worker_filter)
    ws = " WHERE " + " AND ".join(wh) if wh else ""
    totals = conn.execute(f"SELECT COUNT(*) AS r,COUNT(DISTINCT r.worker_id) AS w,COALESCE(SUM(r.quantity),0) AS q,COALESCE(SUM(r.quantity*r.unit_price),0) AS e FROM records r{ws}", pa).fetchone()
    by_worker = conn.execute(f"SELECT w.name AS worker,w.group_name,SUM(r.quantity) AS q,SUM(r.quantity*r.unit_price) AS e FROM records r LEFT JOIN workers w ON r.worker_id=w.id{ws} GROUP BY r.worker_id ORDER BY e DESC", pa).fetchall()
    by_process = conn.execute(f"SELECT p.material,p.process_name,SUM(r.quantity) AS q,SUM(r.quantity*r.unit_price) AS e FROM records r LEFT JOIN processes p ON r.process_id=p.id{ws} GROUP BY r.process_id ORDER BY q DESC", pa).fetchall()
    by_date = conn.execute(f"SELECT r.record_date,SUM(r.quantity) AS q,SUM(r.quantity*r.unit_price) AS e,COUNT(*) AS c FROM records r{ws} GROUP BY r.record_date ORDER BY r.record_date", pa).fetchall()
    conn.close()
    return {'totals':dict(totals),'by_worker':[dict(r) for r in by_worker],'by_process':[dict(r) for r in by_process],'by_date':[dict(r) for r in by_date]}

def list_months():
    conn = get_conn()
    rows = conn.execute("SELECT DISTINCT substr(record_date,1,7) AS m FROM records ORDER BY m DESC").fetchall()
    conn.close()
    return [r['m'] for r in rows]

def _chart_html(stats, title=''):
    t = stats['totals']; w = stats['by_worker']; d = stats['by_date']; p = stats['by_process']
    wj = json.dumps([x['q'] for x in w])
    dj = json.dumps([x['q'] for x in d])
    pj = json.dumps([x['q'] for x in p])
    return f'''<!DOCTYPE html><html><head><meta charset="utf-8"><script src="echarts.min.js"></script>
<style>*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:Microsoft YaHei,sans-serif;background:#f0f2f5;padding:20px}}
.sr{{display:flex;gap:12px;margin-bottom:16px}}
.sc{{flex:1;background:#fff;border-radius:8px;padding:14px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.06)}}
.l{{font-size:12px;color:#888}} .v{{font-size:20px;font-weight:700;color:#2c3e50;margin-top:4px}}
.o{{font-size:11px;color:#27ae60}} .cc{{background:#fff;border-radius:8px;padding:14px;margin-bottom:14px;box-shadow:0 1px 3px rgba(0,0,0,.06)}}
.cb{{width:100%;min-height:220px}} table{{width:100%;border-collapse:collapse;font-size:12px}}
th{{background:#f7f8fa;padding:6px 8px;text-align:left;border-bottom:2px solid #e8e8e8}}
td{{padding:5px 8px;border-bottom:1px solid #f0f0f0}} h3{{font-size:13px;color:#555;margin-bottom:8px}}
</style></head><body>
<h2 style="margin-bottom:12px">{title}</h2>
<div class="sr"><div class="sc"><div class="l">工人数</div><div class="v">{t["w"]}</div></div>
<div class="sc"><div class="l">总产量</div><div class="v">{t["q"]}</div></div>
<div class="sc"><div class="l">总工价</div><div class="v">{YEN}{round(t["e"],2)}</div></div>
<div class="sc"><div class="l">记录数</div><div class="v">{t["r"]}</div></div></div>
<div class="cc"><h3>工人工资排行</h3><div id="w" class="cb"></div></div>
<div class="cc"><h3>日产量趋势</h3><div id="d" class="cb"></div></div>
<div class="cc"><h3>工序产量分布</h3><div id="p" class="cb"></div></div>
<script>
var wc=echarts.init(document.getElementById("w"));wc.setOption({{title:{{show:false}},tooltip:{{trigger:"axis",axisPointer:{{type:"shadow"}}}},xAxis:{{type:"value"}},yAxis:{{type:"category",data:{[x["worker"] for x in w]}}},series:[{{type:"bar",data:{wj},itemStyle:{{color:"#1a73e8",borderRadius:[0,4,4,0]}}}}]}});
var dc=echarts.init(document.getElementById("d"));dc.setOption({{title:{{show:false}},tooltip:{{trigger:"axis"}},xAxis:{{type:"category",data:{[x["record_date"] for x in d]}}},yAxis:{{type:"value"}},series:[{{type:"line",data:{[x["q"] for x in d]},smooth:true,lineStyle:{{color:"#27ae60"}},areaStyle:{{color:"#27ae60",opacity:.15}}}}]}});
var pc=echarts.init(document.getElementById("p"));pc.setOption({{title:{{show:false}},tooltip:{{trigger:"axis",axisPointer:{{type:"shadow"}}}},xAxis:{{type:"value"}},yAxis:{{type:"category",data:{[x["material"]+'-'+x["process_name"] for x in p]}}},series:[{{type:"bar",data:{pj},itemStyle:{{color:"#e67e22",borderRadius:[0,4,4,0]}}}}]}});
</script></body></html>'''

def gen_report(stats, title='生产记录'):
    html = _chart_html(stats, title)
    path = os.path.join(tempfile.gettempdir(), 'report.html')
    with open(path, 'w', encoding='utf-8') as f: f.write(html)
    # Copy echarts to temp
    if os.path.exists(ECHARTS_PATH):
        shutil.copy(ECHARTS_PATH, os.path.join(tempfile.gettempdir(), 'echarts.min.js'))
    webbrowser.open('file://' + path)

def export_excel(stats, filepath, title=''):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        messagebox.showinfo('提示', '请先安装 openpyxl: pip install openpyxl')
        return False
    wb = openpyxl.Workbook(); ws = wb.active
    title = title or '生产记录'
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells('A1:H1')
    t = stats.get('totals', {})
    ws.cell(row=2, column=1, value=f'工人数: {t.get("w",0)}  |  总产量: {t.get("q",0)}  |  总工资: {YEN}{round(t.get("e",0),2)}')
    headers = ['工人', '组别', '件数', '工资']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = Font(bold=True, color='white')
        cell.fill = PatternFill('solid', fgColor='1a73e8')
        cell.alignment = Alignment(horizontal='center')
    for ri, r in enumerate(stats.get('by_worker',[]), 5):
        ws.cell(row=ri, column=1, value=r.get('worker',''))
        ws.cell(row=ri, column=2, value=r.get('group_name',''))
        ws.cell(row=ri, column=3, value=r.get('q',0))
        ws.cell(row=ri, column=4, value=round(r.get('e',0),2))
    wb.save(filepath)
    return True

# ── App ──

class App:
    def __init__(self):
        self.root = Tk()
        self.root.title('生产管理系统')
        self.root.geometry('1200x750')
        self.root.configure(bg=BG)
        self.root.withdraw()
        self.current_user = None
        self.user_perms = {}
        self.show_login()
    
    # ── Permission helpers ──
    def _check_perm(self, perm_key):
        if not self.current_user:
            return False
        if self.current_user.get('role') == 'admin':
            return True
        return bool(self.user_perms.get(perm_key, 0))
    
    def _apply_permissions(self):
        """Enable/disable buttons based on permissions"""
        for widget_name, perm_key in self._perm_widgets:
            for child in self.root.winfo_children():
                self._apply_to_child(child, widget_name, perm_key)
    
    def _apply_to_child(self, parent, widget_name, perm_key):
        if hasattr(parent, '_perm_name') and parent._perm_name == widget_name:
            state = 'normal' if self._check_perm(perm_key) else 'disabled'
            try: parent.config(state=state)
            except: pass
        for child in parent.winfo_children():
            self._apply_to_child(child, widget_name, perm_key)
    
    # ── Login ──
    def show_login(self):
        top = Toplevel(self.root)
        top.title('生产管理系统 登录')
        top.geometry('340x320')
        top.configure(bg=CARD); top.resizable(False, False)
        top.grab_set()
        
        Label(top, text='生产管理系统', font=('Microsoft YaHei',16,'bold'), bg=CARD, fg=PRIMARY).pack(pady=(24,4))
        Label(top, text='请输入账号密码', font=('Microsoft YaHei',9), bg=CARD, fg='#888').pack(pady=(0,12))
        
        f = Frame(top, bg=CARD); f.pack(padx=24)
        Label(f, text='用户名：', bg=CARD, fg='#555', font=('Microsoft YaHei',10)).pack(anchor=W)
        un_entry = Entry(f, width=25, relief='solid', font=('Microsoft YaHei',10), bd=1)
        un_entry.pack(pady=(2,8))
        Label(f, text='密码：', bg=CARD, fg='#555', font=('Microsoft YaHei',10)).pack(anchor=W)
        pw_entry = Entry(f, width=25, relief='solid', font=('Microsoft YaHei',10), show='*', bd=1)
        pw_entry.pack(pady=(2,6))
        
        cred_var = IntVar()
        rf = Frame(f, bg=CARD); rf.pack(fill=X)
        Checkbutton(rf, text='记住密码', variable=cred_var, bg=CARD, font=('Microsoft YaHei',9)).pack(side=LEFT)
        Label(rf, text='默认: admin/admin123', bg=CARD, fg='#999', font=('Microsoft YaHei',8)).pack(side=RIGHT)
        
        def do_login():
            un = un_entry.get().strip(); pw = pw_entry.get()
            if not un or not pw: messagebox.showinfo('提示','请输入用户名和密码'); return
            conn = get_conn()
            r = conn.execute("SELECT * FROM users WHERE username=?", (un,)).fetchone()
            conn.close()
            if not r or not _verify_pw(pw, r['password_hash']):
                messagebox.showerror('错误','用户名或密码错误'); return
            self.current_user = dict(r)
            self.user_perms = get_user_perms(un)
            if cred_var.get(): _save_cred(un, pw)
            else: _clear_cred()
            top.destroy()
            self.root.deiconify()
            self.build_ui()
            self.refresh()
            self._apply_permissions()
        
        Button(f, text='登录', bg=PRIMARY, fg='white', font=('Microsoft YaHei',11,'bold'),
               relief='flat', padx=8, pady=4, cursor='hand2', command=do_login).pack(fill=X, pady=(8,0))
        
        # Load saved credentials
        cred = _load_cred()
        if cred:
            un_entry.insert(0, cred.get('u',''))
            pw_entry.insert(0, cred.get('p',''))
            cred_var.set(1)
        
        un_entry.focus()
        self.root.bind('<Return>', lambda e: do_login())
        self._login_top = top
        self.root.wait_window(top)
    
    # ── Main UI ──
    def build_ui(self):
        # Header
        hdr = Frame(self.root, bg=PRIMARY, height=64); hdr.pack(fill=X); hdr.pack_propagate(False)
        Label(hdr, text='生产记录管理系统', fg='white', bg=PRIMARY, font=('Microsoft YaHei',16,'bold')).pack(padx=24, pady=(10,0), anchor=W)
        Label(hdr, text='工人 · 工序 · 产量 · 工资 · 动态图表', fg='#b0c4de', bg=PRIMARY, font=('Microsoft YaHei',9)).pack(padx=24, pady=(0,8), anchor=W)
        
        main = Frame(self.root, bg=BG); main.pack(fill=BOTH, expand=True, padx=12, pady=(10,8))
        
        # Toolbar
        tb = Frame(main, bg=CARD); tb.pack(fill=X, pady=(0,8))
        self._perm_widgets = []
        
        def make_btn(tb, text, cmd, perm, bg_color):
            btn = Button(tb, text=text, command=cmd, bg=bg_color, fg='white',
                        font=('Microsoft YaHei',9,'bold'), relief='flat', padx=10, pady=3, cursor='hand2')
            btn.pack(side=LEFT, padx=(0,6))
            btn._perm_name = text
            self._perm_widgets.append((text, perm))
            return btn
        
        make_btn(tb, '管理物料', self.manage_materials, 'material_manage', PRIMARY)
        make_btn(tb, '管理工人', self.manage_workers, 'worker_manage', PRIMARY)
        make_btn(tb, '管理工序', self.manage_processes, 'process_manage', PRIMARY)
        make_btn(tb, '汇总查询', self.summary_page, 'summary_view', GREEN)
        # Admin-only buttons - hidden for non-admin
        if self.current_user and self.current_user.get('role') == 'admin':
            make_btn(tb, '用户管理', self.manage_users, 'user_manage', RED)
            make_btn(tb, '权限管理', self.manage_permissions, 'user_manage', GREEN)
        
        # Record entry form
        fm = Frame(main, bg=CARD, highlightbackground='#ddd', highlightthickness=1, padx=14, pady=10)
        fm.pack(fill=X, pady=(0,10))
        Label(fm, text='添加记录', font=('Microsoft YaHei',11,'bold'), bg=CARD, fg=DARK).pack(anchor=W)
        row = Frame(fm, bg=CARD); row.pack(fill=X, pady=(6,0))
        
        fi = Frame(row, bg=CARD); fi.pack(side=LEFT, padx=(0,8))
        Label(fi, text='ID', bg=CARD, fg='#666', font=('Microsoft YaHei',9)).pack(anchor=W, padx=(2,0))
        e_id = Entry(fi, width=6, font=('Microsoft YaHei',11), relief='solid', bd=1, justify=CENTER)
        e_id.pack(); _set_placeholder(e_id, '留空自动')
        
        fw = Frame(row, bg=CARD); fw.pack(side=LEFT, padx=(0,8))
        Label(fw, text='选择工人', bg=CARD, fg='#666', font=('Microsoft YaHei',9)).pack(anchor=W, padx=(2,0))
        cb_worker = ttk.Combobox(fw, width=12, font=('Microsoft YaHei',10), state='readonly'); cb_worker.pack()
        
        fp = Frame(row, bg=CARD); fp.pack(side=LEFT, padx=(0,8))
        Label(fp, text='选择工序', bg=CARD, fg='#666', font=('Microsoft YaHei',9)).pack(anchor=W, padx=(2,0))
        cb_process = ttk.Combobox(fp, width=18, font=('Microsoft YaHei',10), state='readonly'); cb_process.pack()
        
        fq = Frame(row, bg=CARD); fq.pack(side=LEFT, padx=(0,8))
        Label(fq, text='件数', bg=CARD, fg='#666', font=('Microsoft YaHei',9)).pack(anchor=W, padx=(2,0))
        e_qty = Entry(fq, width=8, font=('Microsoft YaHei',11), relief='solid', bd=1, justify=CENTER)
        e_qty.pack()
        
        fd = Frame(row, bg=CARD); fd.pack(side=LEFT, padx=(0,8))
        Label(fd, text='日期', bg=CARD, fg='#666', font=('Microsoft YaHei',9)).pack(anchor=W, padx=(2,0))
        e_date = Entry(fd, width=12, font=('Microsoft YaHei',11), relief='solid', bd=1, justify=CENTER)
        e_date.pack(); e_date.insert(0, date.today().isoformat())
        
        price_label = Label(row, text='', bg=CARD, fg=ACCENT, font=('Microsoft YaHei',10,'bold'))
        price_label.pack(side=LEFT, padx=(4,8))  # Will show unit price
        
        def on_worker_sel(ev):
            sel = cb_worker.current()
            if sel < 0: return
            wid = self._worker_ids[sel] if hasattr(self, '_worker_ids') and sel < len(self._worker_ids) else 0
            if hasattr(self, '_worker_procs') and wid in self._worker_procs:
                procs = self._worker_procs[wid]
                cb_process['values'] = [f"{p['material']} - {p['process_name']}" for p in procs]
            else:
                cb_process['values'] = []
            cb_process.set('')
        
        def on_process_sel(ev):
            sel = cb_process.current()
            if sel < 0: return
            wid = cb_worker.current()
            if wid >= 0 and hasattr(self, '_worker_ids') and wid < len(self._worker_ids):
                widsel = self._worker_ids[wid]
                procs = self._worker_procs.get(widsel, [])
                if sel < len(procs):
                    price_label.config(text=f"单价: {YEN}{procs[sel]['unit_price']}")
        
        cb_worker.bind('<<ComboboxSelected>>', on_worker_sel)
        cb_process.bind('<<ComboboxSelected>>', on_process_sel)
        
        def do_add():
            wid = cb_worker.current()
            pid = cb_process.current()
            qty_s = e_qty.get().strip()
            d = e_date.get().strip()
            if wid < 0 or pid < 0 or not qty_s: messagebox.showinfo('提示','请选择工人、工序并填写件数'); return
            try: qty = float(qty_s)
            except: messagebox.showinfo('提示','件数必须为数字'); return
            widsel = self._worker_ids[wid]
            if widsel in self._worker_procs and pid < len(self._worker_procs[widsel]):
                price = self._worker_procs[widsel][pid]['unit_price']
            else: price = 0
            if add_record(widsel, 0, qty, price, d):
                e_qty.delete(0, END); self.refresh(); messagebox.showinfo('成功','记录已添加')
            else: messagebox.showerror('错误','添加失败')
        
        btn_frame = Frame(row, bg=CARD); btn_frame.pack(side=LEFT)
        Button(btn_frame, text='添加', bg=ACCENT, fg='white', font=('Microsoft YaHei',9,'bold'),
               relief='flat', padx=10, pady=2, cursor='hand2', command=do_add).pack()
        
        # Delete button
        del_frame = Frame(row, bg=CARD); del_frame.pack(side=LEFT, padx=(4,0))
        btn_del = Button(del_frame, text='删除选中', bg=RED, fg='white', font=('Microsoft YaHei',9,'bold'),
                        relief='flat', padx=10, pady=2, cursor='hand2', command=self.delete_selected)
        btn_del.pack()
        btn_del._perm_name = '删除记录'
        self._perm_widgets.append(('删除记录', 'record_delete'))
        
        # Stats bar
        stats_bar = Frame(main, bg='white', relief='solid', bd=1); stats_bar.pack(fill=X, pady=(0,6))
        self._stat_labels = {}
        for key, label in [('w','工人数'),('q','总产量'),('e','总工价'),('r','记录数')]:
            cell = Frame(stats_bar, bg='white'); cell.pack(side=LEFT, expand=True, padx=8, pady=6)
            Label(cell, text=label, bg='white', fg='#888', font=('Microsoft YaHei',8)).pack()
            self._stat_labels[key] = Label(cell, text='0', bg='white', fg=DARK, font=('Microsoft YaHei',14,'bold'))
            self._stat_labels[key].pack()
        
        # Tree table
        tree_frame = Frame(main); tree_frame.pack(fill=BOTH, expand=True)
        tr = ttk.Treeview(tree_frame, columns=('id','material','process','worker','group','qty','price','wage','date'),
                          show='headings', height=12)
        for col, text, w in [('id','ID',40),('material','物料',80),('process','工序',100),('worker','姓名',70),
                            ('group','组别',70),('qty','件数',60),('price','单价',60),('wage','工资',70),('date','日期',100)]:
            tr.heading(col, text=text); tr.column(col, width=w, anchor='center' if col in ('id','qty','price','wage') else 'w')
        tr.pack(side=LEFT, fill=BOTH, expand=True)
        vsb = Scrollbar(tree_frame, orient=VERTICAL, command=tr.yview); tr.configure(yscrollcommand=vsb.set)
        vsb.pack(side=RIGHT, fill=Y)
        self._tree = tr
        
        # Bottom buttons
        bottom = Frame(main, bg=BG); bottom.pack(fill=X, pady=(4,0))
        Button(bottom, text='📊 生成图表报告', command=self.open_chart, bg=PRIMARY, fg='white',
               font=('Microsoft YaHei',9,'bold'), relief='flat', padx=10, cursor='hand2').pack(side=LEFT, padx=(0,6))
        Button(bottom, text='📅 月度汇总', command=self.monthly_summary, bg=GREEN, fg='white',
               font=('Microsoft YaHei',9,'bold'), relief='flat', padx=10, cursor='hand2').pack(side=LEFT)
    
    def refresh(self):
        # Load workers & processes for form
        self._worker_ids = []
        self._worker_procs = {}
        for w in get_workers():
            self._worker_ids.append(w['id'])
            procs = []
            for p in get_processes():
                wp = get_worker_processes(w['id'])
                if p['id'] in wp or True:  # Show all processes for now
                    procs.append(p)
            self._worker_procs[w['id']] = procs
        
        # Update combobox
        for child in self.root.winfo_children():
            self._find_and_update(child)
        
        # Refresh table
        self._refresh_table()
    
    def _find_and_update(self, parent):
        if isinstance(parent, ttk.Combobox) and parent.winfo_children():
            pass  # Will be handled by the main loop
        for child in parent.winfo_children():
            self._find_and_update(child)
    
    def _refresh_table(self):
        tr = self._tree
        tr.delete(*tr.get_children())
        records = get_all_records(self.current_user)
        for r in records[:200]:
            wage = r['quantity'] * r['unit_price']
            tr.insert('', END, values=(r['id'],r['material'],r['process_name'],r['worker_name'],
                                       r['group_name'],r['quantity'],f'{YEN}{r["unit_price"]}',
                                       f'{YEN}{round(wage,2)}',r['record_date']))
    
    def delete_selected(self):
        sel = self._tree.selection()
        if not sel: return
        if not messagebox.askyesno('确认','确定删除选中的记录？'): return
        for item in sel:
            vals = self._tree.item(item, 'values')
            if vals: delete_record(int(vals[0]))
        self.refresh()
    
    def open_chart(self):
        stats = get_stats()
        gen_report(stats, '生产记录统计')
    
    def monthly_summary(self):
        top = Toplevel(self.root); top.title('月度汇总'); top.geometry('700x500')
        top.configure(bg=CARD); top.grab_set()
        Label(top, text='选择月份', font=('Microsoft YaHei',12,'bold'), bg=CARD, fg=DARK).pack(pady=(10,4))
        months = list_months()
        if not months:
            Label(top, text='暂无数据', bg=CARD, fg='#888').pack(); return
        cb = ttk.Combobox(top, values=months, state='readonly', width=12, font=('Microsoft YaHei',11))
        cb.pack(pady=4); cb.set(months[0] if months else '')
        
        def do():
            m = cb.get()
            if not m: return
            stats = get_stats(start_date=m+'-01', end_date=m+'-31')
            gen_report(stats, f'{m} 月度汇总')
        
        Button(top, text='生成报告', command=do, bg=PRIMARY, fg='white',
               font=('Microsoft YaHei',10,'bold'), relief='flat', padx=14, pady=4, cursor='hand2').pack(pady=6)
        
        tr = ttk.Treeview(top, columns=('name','group','qty','wage'), show='headings', height=10)
        for col, text, w in [('name','工人',100),('group','组别',100),('qty','件数',80),('wage','工资',100)]:
            tr.heading(col, text=text); tr.column(col, width=w)
        tr.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        def refresh():
            tr.delete(*tr.get_children())
            m = cb.get()
            if not m: return
            stats = get_stats(start_date=m+'-01', end_date=m+'-31')
            for r in stats.get('by_worker',[]):
                tr.insert('', END, values=(r['worker'],r['group_name'],r['q'],round(r['e'],2)))
        
        cb.bind('<<ComboboxSelected>>', lambda e: refresh())
        if months: refresh()
        
        def do_export():
            m = cb.get()
            if not m: return
            path = os.path.join(_BASE, f'月度汇总_{m}.xlsx')
            stats = get_stats(start_date=m+'-01', end_date=m+'-31')
            if export_excel(stats, path, f'{m} 月度汇总'):
                messagebox.showinfo('成功', f'已导出: {path}')
        
        Button(top, text='导出 Excel', command=do_export, bg=GREEN, fg='white',
               font=('Microsoft YaHei',9,'bold'), relief='flat', padx=10, cursor='hand2').pack(pady=(0,8))
    
    # ── Management dialogs ──
    def manage_materials(self):
        top = Toplevel(self.root); top.title('管理物料'); top.geometry('500x400')
        top.configure(bg=CARD); top.grab_set()
        Label(top, text='物料列表', font=('Microsoft YaHei',12,'bold'), bg=CARD, fg=DARK).pack(anchor=W, padx=16, pady=(10,4))
        tr = ttk.Treeview(top, columns=('name','price'), show='headings', height=10)
        tr.heading('name', text='名称'); tr.heading('price', text='单价(元)')
        tr.column('name', width=200); tr.column('price', width=100)
        def ref():
            tr.delete(*tr.get_children())
            for m in get_materials(): tr.insert('', END, values=(m['name'], m['price']))
        ref(); tr.pack(fill=BOTH, expand=True, padx=16)
        f = Frame(top, bg=CARD); f.pack(fill=X, padx=16, pady=6)
        e_n = Entry(f, width=12, font=('Microsoft YaHei',11), relief='solid', bd=1)
        e_n.pack(side=LEFT, padx=(0,4)); _set_placeholder(e_n, '名称')
        e_p = Entry(f, width=8, font=('Microsoft YaHei',11), relief='solid', bd=1)
        e_p.pack(side=LEFT, padx=(0,4)); _set_placeholder(e_p, '单价')
        def do_add():
            n = e_n.get().strip(); p = e_p.get().strip()
            if n == getattr(e_n,'_ph_text',None): n = ''
            if not n: return
            try: price = float(p) if p else 0
            except: price = 0
            if add_material(n, price): ref(); e_n.delete(0,END); e_p.delete(0,END)
        Button(f, text='添加', bg=ACCENT, fg='white', font=('Microsoft YaHei',9,'bold'),
               relief='flat', padx=8, cursor='hand2', command=do_add).pack(side=LEFT)
        def do_del():
            sel = tr.selection()
            if sel:
                v = tr.item(sel[0],'values')
                if messagebox.askyesno('确认', f'删除物料 "{v[0]}"？'): delete_material(int(v[0])); ref()
        Button(f, text='删除', bg=RED, fg='white', relief='flat', padx=8, cursor='hand2', command=do_del).pack(side=LEFT, padx=(4,0))
    
    def manage_workers(self):
        top = Toplevel(self.root); top.title('管理工人'); top.geometry('620x500')
        top.configure(bg=CARD); top.grab_set()
        Label(top, text='工人列表', font=('Microsoft YaHei',12,'bold'), bg=CARD, fg=DARK).pack(anchor=W, padx=16, pady=(10,4))
        tr = ttk.Treeview(top, columns=('name','group'), show='headings', height=8)
        tr.heading('name', text='姓名'); tr.heading('group', text='组别')
        tr.column('name', width=180); tr.column('group', width=200)
        def ref():
            tr.delete(*tr.get_children())
            for w in get_workers(): tr.insert('', END, values=(w['name'], w['group_name']))
        ref(); tr.pack(padx=16, fill=BOTH, expand=True, pady=(4,4))
        top._tr = tr
        f = Frame(top, bg=CARD); f.pack(side=BOTTOM, fill=X, padx=16, pady=(0,6))
        e_n = Entry(f, width=12, font=('Microsoft YaHei',11), relief='solid', bd=1); e_n.pack(side=LEFT, padx=(0,4))
        _set_placeholder(e_n, '姓名')
        e_g = Entry(f, width=12, font=('Microsoft YaHei',11), relief='solid', bd=1); e_g.pack(side=LEFT, padx=(0,4))
        _set_placeholder(e_g, '组别')
        def do_add():
            n = e_n.get().strip(); g = e_g.get().strip()
            if n == getattr(e_n,'_ph_text',None): n = ''
            if not n: return
            if add_worker(n, g): ref(); e_n.delete(0,END); e_g.delete(0,END)
        Button(f, text='添加', bg=ACCENT, fg='white', font=('Microsoft YaHei',9,'bold'),
               relief='flat', padx=8, cursor='hand2', command=do_add).pack(side=LEFT, padx=(4,0))
        def do_del():
            sel = tr.selection()
            if sel:
                v = tr.item(sel[0],'values')
                if messagebox.askyesno('确认', f'删除工人 "{v[0]}"？'):
                    delete_worker(int(tr.item(sel[0],'values')[0])); ref()
        Button(f, text='删除', bg=RED, fg='white', relief='flat', padx=8, cursor='hand2', command=do_del).pack(side=LEFT, padx=(4,0))
    
    def manage_processes(self):
        top = Toplevel(self.root); top.title('管理工序'); top.geometry('600x450')
        top.configure(bg=CARD); top.grab_set()
        Label(top, text='工序列表', font=('Microsoft YaHei',12,'bold'), bg=CARD, fg=DARK).pack(anchor=W, padx=16, pady=(10,4))
        tr = ttk.Treeview(top, columns=('material','process','price'), show='headings', height=10)
        tr.heading('material', text='物料'); tr.heading('process', text='工序'); tr.heading('price', text='单价')
        tr.column('material', width=130); tr.column('process', width=200); tr.column('price', width=80)
        def ref():
            tr.delete(*tr.get_children())
            for p in get_processes(): tr.insert('', END, values=(p['material'], p['process_name'], p['unit_price']))
        ref(); tr.pack(fill=BOTH, expand=True, padx=16)
        f = Frame(top, bg=CARD); f.pack(fill=X, padx=16, pady=6)
        e_m = Entry(f, width=10, font=('Microsoft YaHei',11), relief='solid', bd=1); e_m.pack(side=LEFT, padx=(0,4))
        _set_placeholder(e_m, '物料'); e_pn = Entry(f, width=12, font=('Microsoft YaHei',11), relief='solid', bd=1)
        e_pn.pack(side=LEFT, padx=(0,4)); _set_placeholder(e_pn, '工序')
        e_pr = Entry(f, width=6, font=('Microsoft YaHei',11), relief='solid', bd=1); e_pr.pack(side=LEFT, padx=(0,4))
        _set_placeholder(e_pr, '单价')
        def do_add():
            m = e_m.get().strip(); pn = e_pn.get().strip(); pr = e_pr.get().strip()
            if m == getattr(e_m,'_ph_text',None): m = ''
            if not m or not pn: return
            try: price = float(pr) if pr else 0
            except: price = 0
            if add_process(m, pn, price): ref(); e_m.delete(0,END); e_pn.delete(0,END); e_pr.delete(0,END)
        Button(f, text='添加', bg=ACCENT, fg='white', font=('Microsoft YaHei',9,'bold'),
               relief='flat', padx=8, cursor='hand2', command=do_add).pack(side=LEFT)
        def do_del():
            sel = tr.selection()
            if sel:
                v = tr.item(sel[0],'values')
                if messagebox.askyesno('确认', f'删除工序 "{v[1]}"？'): delete_process(int(tr.item(sel[0],'values')[0])); ref()
        Button(f, text='删除', bg=RED, fg='white', relief='flat', padx=8, cursor='hand2', command=do_del).pack(side=LEFT, padx=(4,0))
    
    def manage_users(self):
        top = Toplevel(self.root); top.title('用户管理'); top.geometry('700x500')
        top.configure(bg=CARD); top.grab_set()
        Label(top, text='用户列表', font=('Microsoft YaHei',12,'bold'), bg=CARD, fg=DARK).pack(anchor=W, padx=16, pady=(10,4))
        tr = ttk.Treeview(top, columns=('username','display','role','created'), show='headings', height=10)
        tr.heading('username', text='用户名'); tr.heading('display', text='显示名')
        tr.heading('role', text='角色'); tr.heading('created', text='创建时间')
        tr.column('username', width=120); tr.column('display', width=120)
        tr.column('role', width=100); tr.column('created', width=150)
        def ref():
            tr.delete(*tr.get_children())
            for u in get_users():
                role_text = {'admin':'管理员','leader':'班组长','worker':'普通用户'}.get(u['role'], u['role'])
                tr.insert('', END, values=(u['username'], u['display_name'], role_text, u.get('created_at','')))
        ref(); tr.pack(fill=BOTH, expand=True, padx=16, pady=(0,4))
        f = Frame(top, bg=CARD); f.pack(fill=X, padx=16, pady=(0,6))
        e_un = Entry(f, width=12, font=('Microsoft YaHei',11), relief='solid', bd=1); e_un.pack(side=LEFT, padx=(0,4))
        _set_placeholder(e_un, '用户名')
        e_dn = Entry(f, width=10, font=('Microsoft YaHei',11), relief='solid', bd=1); e_dn.pack(side=LEFT, padx=(0,4))
        _set_placeholder(e_dn, '显示名')
        e_pw = Entry(f, width=10, font=('Microsoft YaHei',11), relief='solid', bd=1, show='*')
        e_pw.pack(side=LEFT, padx=(0,4)); _set_placeholder(e_pw, '密码')
        role_cb = ttk.Combobox(f, values=['管理员','班组长','普通用户'], state='readonly', width=8)
        role_cb.pack(side=LEFT, padx=(0,4)); role_cb.set('普通用户')
        def do_add():
            un = e_un.get().strip(); dn = e_dn.get().strip(); pw = e_pw.get().strip()
            if un == getattr(e_un,'_ph_text',None): un = ''
            if dn == getattr(e_dn,'_ph_text',None): dn = ''
            if pw == getattr(e_pw,'_ph_text',None): pw = ''
            if not un or not pw: messagebox.showinfo('提示','用户名和密码不能为空'); return
            role_map = {'管理员':'admin','班组长':'leader','普通用户':'worker'}
            role = role_map.get(role_cb.get(), 'worker')
            if add_user(un, pw, dn, role):
                ref(); e_un.delete(0,END); e_dn.delete(0,END); e_pw.delete(0,END)
            else: messagebox.showinfo('提示','用户名已存在')
        Button(f, text='添加', bg=ACCENT, fg='white', font=('Microsoft YaHei',9,'bold'),
               relief='flat', padx=8, cursor='hand2', command=do_add).pack(side=LEFT, padx=(4,0))
        def do_del():
            sel = tr.selection()
            if not sel: return
            v = tr.item(sel[0],'values')
            if v[0] == 'admin': messagebox.showinfo('提示','不能删除 admin 账户'); return
            if messagebox.askyesno('确认', f'删除用户 "{v[0]}"？'):
                delete_user(v[0]); ref()
        Button(f, text='删除', bg=RED, fg='white', font=('Microsoft YaHei',9,'bold'),
               relief='flat', padx=8, cursor='hand2', command=do_del).pack(side=LEFT, padx=(4,0))
    
    def manage_permissions(self):
        top = Toplevel(self.root); top.title('权限管理'); top.geometry('800x500')
        top.configure(bg=CARD); top.grab_set()
        Label(top, text='权限管理', font=('Microsoft YaHei',12,'bold'), bg=CARD, fg=DARK).pack(anchor=W, padx=16, pady=(10,4))
        Label(top, text='点击复选框切换权限', font=('Microsoft YaHei',9), bg=CARD, fg='#888').pack(anchor=W, padx=16)
        
        frame = Frame(top, bg=CARD)
        frame.pack(fill=BOTH, expand=True, padx=16, pady=(4,8))
        
        users = get_users()
        # Header row
        Label(frame, text='用户', bg=CARD, fg=DARK, font=('Microsoft YaHei',9,'bold'), width=10, anchor='w').grid(row=0, column=0, padx=2, pady=2, sticky='w')
        perm_labels = [('record_add','添加记录'),('record_delete','删除记录'),('record_edit','编辑记录'),
                      ('material_manage','管理物料'),('worker_manage','管理工人'),('process_manage','管理工序'),
                      ('assignment_manage','工序分配'),('chart_view','查看图表'),('summary_view','查看汇总'),
                      ('export_excel','导出Excel'),('user_manage','用户管理')]
        perms = [p[0] for p in perm_labels]
        for ci, (pk, pl) in enumerate(perm_labels, 1):
            Label(frame, text=pl, bg=CARD, fg='#555', font=('Microsoft YaHei',8), width=8, anchor='w').grid(row=0, column=ci, padx=1)
        
        self._perm_vars = {}
        for ri, u in enumerate(users, 1):
            un = u['username']
            Label(frame, text=un, bg=CARD, fg=DARK, font=('Microsoft YaHei',9), width=10, anchor='w').grid(row=ri, column=0, padx=2, pady=1, sticky='w')
            perms_dict = get_user_perms(un)
            for ci, pk in enumerate(perms, 1):
                var = IntVar(value=perms_dict.get(pk, 0))
                self._perm_vars[(un, pk)] = var
                def make_cmd(u_name, p_key):
                    return lambda: set_user_perm(u_name, p_key, self._perm_vars[(u_name, p_key)].get())
                cb = Checkbutton(frame, variable=var, bg=CARD, command=make_cmd(un, pk))
                cb.grid(row=ri, column=ci)
    
    def summary_page(self):
        top = Toplevel(self.root); top.title('汇总查询'); top.geometry('800x550')
        top.configure(bg=CARD); top.grab_set()
        Label(top, text='汇总查询', font=('Microsoft YaHei',12,'bold'), bg=CARD, fg=DARK).pack(anchor=W, padx=16, pady=(10,4))
        
        # Filters
        ff = Frame(top, bg=CARD); ff.pack(fill=X, padx=16)
        Label(ff, text='起始日期:', bg=CARD, fg='#555', font=('Microsoft YaHei',9)).pack(side=LEFT)
        e_start = Entry(ff, width=12, font=('Microsoft YaHei',10), relief='solid', bd=1)
        e_start.pack(side=LEFT, padx=(2,8))
        Label(ff, text='结束日期:', bg=CARD, fg='#555', font=('Microsoft YaHei',9)).pack(side=LEFT)
        e_end = Entry(ff, width=12, font=('Microsoft YaHei',10), relief='solid', bd=1)
        e_end.pack(side=LEFT, padx=(2,8))
        
        workers = get_workers()
        cb_worker = ttk.Combobox(ff, values=['全部']+[w['name'] for w in workers], state='readonly', width=10)
        cb_worker.pack(side=LEFT, padx=(4,4)); cb_worker.set('全部')
        
        def do_query():
            sd = e_start.get().strip() or None
            ed = e_end.get().strip() or None
            wf = None
            sel = cb_worker.current()
            if sel > 0 and sel <= len(workers):
                wf = workers[sel-1]['id']
            stats = get_stats(start_date=sd, end_date=ed, worker_filter=wf)
            gen_report(stats, '汇总查询结果')
        
        Button(ff, text='查询', command=do_query, bg=PRIMARY, fg='white',
               font=('Microsoft YaHei',9,'bold'), relief='flat', padx=10, cursor='hand2').pack(side=LEFT, padx=(8,0))
        
        # Stats table
        tr = ttk.Treeview(top, columns=('worker','group','qty','wage'), show='headings', height=12)
        tr.heading('worker', text='工人'); tr.heading('group', text='组别')
        tr.heading('qty', text='件数'); tr.heading('wage', text='工资')
        tr.column('worker', width=120); tr.column('group', width=120)
        tr.column('qty', width=100); tr.column('wage', width=120)
        tr.pack(fill=BOTH, expand=True, padx=16, pady=(8,4))
        
        def do_query_table():
            sd = e_start.get().strip() or None
            ed = e_end.get().strip() or None
            wf = None
            sel = cb_worker.current()
            if sel > 0 and sel <= len(workers):
                wf = workers[sel-1]['id']
            stats = get_stats(start_date=sd, end_date=ed, worker_filter=wf)
            tr.delete(*tr.get_children())
            for r in stats.get('by_worker',[]):
                tr.insert('', END, values=(r['worker'], r['group_name'], r['q'], round(r['e'],2)))
        
        def do_export():
            sd = e_start.get().strip() or None
            ed = e_end.get().strip() or None
            wf = None
            sel = cb_worker.current()
            if sel > 0 and sel <= len(workers):
                wf = workers[sel-1]['id']
            stats = get_stats(start_date=sd, end_date=ed, worker_filter=wf)
            path = os.path.join(_BASE, '汇总查询.xlsx')
            title = '汇总查询'
            if sd or ed: title += f' ({sd or ""}~{ed or ""})'
            if export_excel(stats, path, title):
                messagebox.showinfo('成功', f'已导出: {path}')
        
        bf = Frame(top, bg=CARD); bf.pack(fill=X, padx=16, pady=(0,8))
        Button(bf, text='查询表格', command=do_query_table, bg=PRIMARY, fg='white',
               font=('Microsoft YaHei',9,'bold'), relief='flat', padx=10, cursor='hand2').pack(side=LEFT, padx=(0,6))
        Button(bf, text='导出 Excel', command=do_export, bg=GREEN, fg='white',
               font=('Microsoft YaHei',9,'bold'), relief='flat', padx=10, cursor='hand2').pack(side=LEFT)

if __name__ == '__main__':
    init_db()
    app = App()
    app.root.mainloop()
