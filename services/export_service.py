"""Excel 导出服务"""
import os
from config import YEN
from utils.logger import logger


def export_excel(stats, filepath, title='', records=None):
    """导出统计数据到 Excel，records 为可选的生产记录明细"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        from tkinter import messagebox
        messagebox.showinfo('\u63d0\u793a', '\u8bf7\u5148\u5b89\u88c5 openpyxl: pip install openpyxl')
        return False
    wb = openpyxl.Workbook()
    ws = wb.active
    title_str = title or '\u751f\u4ea7\u8bb0\u5f55'
    ws.cell(row=1, column=1, value=title_str).font = Font(bold=True, size=13)
    ws.merge_cells('A1:H1')
    t = stats.get('totals', {})
    ws.cell(row=2, column=1, value=f'\u5de5\u4eba\u6570: {t.get("w",0)}  |  \u603b\u4ea7\u91cf: {t.get("q",0)}  |  \u603b\u5de5\u4ef7: {YEN}{round(t.get("e",0),2)}')
    headers = ['\u5de5\u4eba', '\u7ec4\u522b', '\u4ef6\u6570', '\u5de5\u4ef7']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = Font(bold=True, color='00ffffff')
        cell.fill = PatternFill('solid', fgColor='001a73e8')
        cell.alignment = Alignment(horizontal='center')
    for ri, r in enumerate(stats.get('by_worker', []), 5):
        ws.cell(row=ri, column=1, value=r.get('worker', ''))
        ws.cell(row=ri, column=2, value=r.get('group_name', ''))
        ws.cell(row=ri, column=3, value=r.get('q', 0))
        ws.cell(row=ri, column=4, value=round(r.get('e', 0), 2))
    if records:
        # 写入生产记录明细
        detail_row = len(stats.get('by_worker', [])) + 7
        ws.cell(row=detail_row, column=1, value='生产记录明细').font = Font(bold=True, size=12)
        ws.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=9)
        detail_row += 1
        det_headers = ['ID', '物料编号', '工序', '工人', '班组', '件数', '工价', '总价', '日期']
        for ci, h in enumerate(det_headers, 1):
            cell = ws.cell(row=detail_row, column=ci, value=h)
            cell.font = Font(bold=True, color='00ffffff')
            cell.fill = PatternFill('solid', fgColor='002980b9')
            cell.alignment = Alignment(horizontal='center')
        for ri, r in enumerate(records, detail_row + 1):
            ws.cell(row=ri, column=1, value=r.get('id', ''))
            ws.cell(row=ri, column=2, value=r.get('material_code', ''))
            ws.cell(row=ri, column=3, value=r.get('process_name', ''))
            ws.cell(row=ri, column=4, value=r.get('worker_name', ''))
            ws.cell(row=ri, column=5, value=r.get('group_name', ''))
            ws.cell(row=ri, column=6, value=r.get('quantity', 0))
            ws.cell(row=ri, column=7, value=r.get('unit_price', 0))
            ws.cell(row=ri, column=8, value=round(r.get('quantity', 0) * r.get('unit_price', 0), 2))
            ws.cell(row=ri, column=9, value=r.get('record_date', ''))
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 14
    wb.save(filepath)
    logger.info(f'Excel \u5df2\u5bfc\u51fa: {filepath}')
    return True
